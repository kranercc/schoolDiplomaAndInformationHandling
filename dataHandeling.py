#! python3
import openpyxl,time,os
locuFisierului = "C:\\Users\\INFO\\Desktop\\diplome\\test.xlsx"
wb = openpyxl.load_workbook(locuFisierului)
sheet = wb.get_sheet_by_name("Sheet1")
os.system("cls")
nume = input("Nume - Initiala Tatalui - Prenume (fara \"-\"): ")
def getNames():
    global nume
    for i in range(2,100000):
        a=sheet.cell(row=i,column=1).value
        rownr = i
        date = ["Nume: ","Anul absolvirii: ","Ce a absolvit: ","Calificare Profesionala: ","Ridicat Diploma: "] # 5 nume 6 an 7 forma 8 calificare 9 diploma 
        if a == nume:
            for coloana in range(1,6):
                a=sheet.cell(row=rownr,column=coloana).value
                date.append(a)
            if date[5] == None:
                print("Informatia despre nume nu a fost adaugata")
            else:
                print(date[0],date[5])
            if date[6] == None:
                print("Informatia despre calificarea profesionala nu a fost adaugata")
            else:
                print(date[1],date[6])
            if date[7] == None:
                print("Informatia despre forma absolvita [liceu/sc maistri etc] nu a fost adaugata")
            else:
                print(date[2],date[7])
            if date[8] == None:
                print("Informatia despre calificare nu a fost adaugata")
            else:
                print(date[3],date[8])
            if date[9] == None:
                print("Informatia despre diploma nu a fost adaugata")
            else:
                print(date[4],date[9])
            break
        if a == None:
            break
    return a

def changeDiploma():
    global nume,wb
    for i in range(1,100000):
        a = sheet.cell(row=i,column=1).value
        rownr = i
        if a == nume:
            sheet.cell(row=rownr,column=5).value = "DA"
            wb.save("test.xlsx")
            break
actions = input("1. Verificari date\n2. I-a fost data diploma?: ")
if actions == "1":
    getNames()
elif actions == "2":
    safety = input("Sunteti sigur de schimbarea celulei in DA?\n Daca DA atunci scrieti DA: ")
    if safety == "DA":
        changeDiploma()
        print("Modificarile au fost facute.")
    else:
        pass

time.sleep(60)
